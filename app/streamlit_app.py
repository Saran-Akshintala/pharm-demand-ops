"""
Streamlit app for pharmacy demand forecasting.
Upload Excel files, get predictions, and download results.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import json
import sys
from io import BytesIO

# Add current directory to path for imports
sys.path.append('.')
from utils import (
    preprocess_sales_features, 
    prepare_features_for_prediction,
    load_model,
    parse_order_scheme,
    reconstruct_order_prediction,
    validate_input_data,
    apply_predicted_order_business_rules,
    compute_predicted_order_with_adjustments
)
from retrain import ModelRetrainer, ModelRegistry
from enhanced_grid import create_enhanced_grid, process_grid_changes, show_grid_changes_summary

# Load custom CSS (deprecated) - default Streamlit styling only
def load_css():
    """Deprecated: No custom CSS injection. Using Streamlit defaults."""
    pass

def inject_pandas_table_css():
    """Deprecated: No CSS injection. Using Streamlit defaults."""
    pass

def render_dataframe(df: pd.DataFrame):
    """Render a dataframe using Streamlit's default styling and viewer."""
    st.dataframe(df, use_container_width=True)

# Page configuration
st.set_page_config(
    page_title="Pharmacy Demand Forecasting",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Default Streamlit styling only (no CSS injection)

def process_uploaded_file(uploaded_file, apply_box=True, box_tolerance=2, apply_scm=True, scm_tolerance=2):
    """Process uploaded Excel file and return predictions with configurable adjustments"""
    try:
        # Read the uploaded file
        df = pd.read_excel(uploaded_file)
        
        # Validate input data
        validation_results = validate_input_data(df)
        
        if validation_results['errors']:
            st.error("❌ Data validation failed:")
            for error in validation_results['errors']:
                st.error(f"• {error}")
            return None, None, None
        
        if validation_results['warnings']:
            st.warning("⚠️ Data validation warnings:")
            for warning in validation_results['warnings']:
                st.warning(f"• {warning}")
        
        # Load model and make predictions
        script_dir = Path(__file__).parent.absolute()
        
        # Always use absolute path from script location
        model_path = script_dir.parent / "models" / "order_predictor.pkl"
        
        
        if not model_path.exists():
            st.error("❌ Model not found. Please train the model first.")
            return None, None, None
        
        model = load_model(str(model_path))
        
        # Load model info for feature columns
        model_info_path = script_dir.parent / "models" / "model_info.json"
            
        if model_info_path.exists():
            with open(model_info_path) as f:
                model_info = json.load(f)
            feature_columns = model_info['feature_columns']
        else:
            # Fallback feature columns
            feature_columns = ['L7', 'L15', 'L30', 'L45', 'L60', 'L75', 'L90']
        
        # Preprocess features
        df_processed = preprocess_sales_features(df)
        
        # Prepare features for prediction
        X = prepare_features_for_prediction(df_processed, feature_columns)
        
        # Make predictions
        predictions = model.predict(X)
        
        # Add predictions to original dataframe
        df_result = df.copy()
        df_result['Predicted_Base_Quantity'] = predictions.round().astype(int)
        
        # Use new deterministic logic to compute Predicted_Order from Predicted_Base_Quantity
        df_result = compute_predicted_order_with_adjustments(
            df_result, 
            apply_box=apply_box, 
            box_tolerance=box_tolerance,
            apply_scm=apply_scm, 
            scm_tolerance=scm_tolerance
        )
        
        # Apply business rules to Predicted_Order column
        df_result, styling_info = apply_predicted_order_business_rules(df_result)
        
        # Apply expiry highlighting
        from utils import apply_expiry_highlighting
        expiry_styling = apply_expiry_highlighting(df_result)
        
        # Reorder columns to place predictions next to Order column (fix from memory)
        order_col = None
        # Check for various order column name variations
        for col in df_result.columns:
            col_lower = col.lower()
            if col_lower in ['order', 'ord', 'oreder']:
                order_col = col
                break
        
        if order_col:
            # Use df_result.columns (after adding predictions) instead of df.columns
            cols = list(df_result.columns)
            
            # Remove prediction columns from their current positions
            cols = [col for col in cols if col not in ['Predicted_Order', 'Predicted_Base_Quantity']]
            
            # Find the position of the Order column
            order_idx = cols.index(order_col)
            
            # Insert prediction columns right after Order column
            cols.insert(order_idx + 1, 'Predicted_Order')
            cols.insert(order_idx + 2, 'Predicted_Base_Quantity')
            
            # Reorder the dataframe
            df_result = df_result.reindex(columns=cols)
        
        return df_result, df_processed, styling_info, expiry_styling
        
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        return None, None, None, None

def create_visualizations(df_result, df_processed):
    """Create visualizations for the predictions"""
    
    # Prediction distribution
    fig_dist = px.histogram(
        df_result, 
        x='Predicted_Base_Quantity',
        title="📊 Distribution of Predicted Order Quantities",
        nbins=30,
        color_discrete_sequence=['#667eea']
    )
    fig_dist.update_layout(
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font_color='white'
    )
    st.plotly_chart(fig_dist, use_container_width=True)
    
    # Top products by predicted quantity
    if len(df_result) > 0:
        top_products = df_result.nlargest(10, 'Predicted_Base_Quantity')[['Name', 'Predicted_Base_Quantity', 'Predicted_Order']]
        
        fig_top = px.bar(
            top_products,
            x='Predicted_Base_Quantity',
            y='Name',
            orientation='h',
            title="🔝 Top 10 Products by Predicted Quantity",
            color='Predicted_Base_Quantity',
            color_continuous_scale='viridis'
        )
        fig_top.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font_color='white',
            height=500
        )
        st.plotly_chart(fig_top, use_container_width=True)

def main():
    """Main application function"""
    
    # App header with modern styling
    st.markdown('<div class="main-header">', unsafe_allow_html=True)
    st.markdown("# 💊 Pharmacy Demand Forecasting")
    st.markdown("### AI-Powered Order Prediction System")
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Add section divider
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    
    # Create tabs for different functionalities
    tab1, tab2 = st.tabs(["📊 Demand Prediction", "🔄 Model Retraining"])
    
    with tab1:
        prediction_tab()
    
    with tab2:
        retraining_tab()


def prediction_tab():
    """Original prediction functionality in a tab"""
    
    # Sidebar with model information
    with st.sidebar:
        st.markdown("## 🤖 Model Information")
        
        # Try to load model info
        try:
            script_dir = Path(__file__).parent.absolute()
            model_info_path = script_dir.parent / "models" / "model_info.json"
                
            if model_info_path.exists():
                with open(model_info_path) as f:
                    model_info = json.load(f)
                
                st.markdown("### 📊 Performance Metrics")
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("RMSE", f"{model_info['performance']['rmse']:.3f}")
                    st.metric("R²", f"{model_info['performance']['r2']:.3f}")
                with col2:
                    st.metric("MAE", f"{model_info['performance']['mae']:.3f}")
                    st.metric("Features", len(model_info['feature_columns']))
                
                st.markdown("### 🎯 Model Type")
                st.info(f"**{model_info['model_type']}**")
                
                st.markdown("### 📈 Training Data")
                st.write(f"Training samples: {model_info['training_data_size']:,}")
                st.write(f"Test samples: {model_info['test_data_size']:,}")
            else:
                st.warning("Model info not found. Please train the model first.")
        except Exception as e:
            st.error(f"Error loading model info: {e}")
    
    # File upload section
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    st.markdown("## 📁 Upload Pharmacy Data")
    
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload your pharmacy order data in Excel format"
    )
    st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_file is not None:
        st.markdown('<div class="results-section">', unsafe_allow_html=True)
        
        # Process the file
        with st.spinner("🔄 Processing your data..."):
            result = process_uploaded_file(uploaded_file)
            if result[0] is not None:
                df_result, df_processed, styling_info, expiry_styling = result
            else:
                df_result, df_processed, styling_info, expiry_styling = None, None, None, None
        
        if df_result is not None:
            st.success("✅ Predictions generated successfully!")
            
            # Display summary statistics
            st.markdown("## 📈 Prediction Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Products", len(df_result))
            with col2:
                avg_prediction = df_result['Predicted_Base_Quantity'].mean()
                st.metric("Avg Predicted Qty", f"{avg_prediction:.1f}")
            with col3:
                max_prediction = df_result['Predicted_Base_Quantity'].max()
                st.metric("Max Predicted Qty", f"{max_prediction}")
            with col4:
                total_prediction = df_result['Predicted_Base_Quantity'].sum()
                st.metric("Total Predicted Qty", f"{total_prediction:,}")
            
            # Add section divider
            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
            
            # Visualizations
            st.markdown("## 📊 Prediction Analytics")
            create_visualizations(df_result, df_processed)
            
            # Add section divider
            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
            
            # Preview table with business rule styling
            st.markdown("## 📋 Detailed Results (All Rows)")
            
            # Add UI controls for box and scheme adjustments
            st.markdown("### ⚙️ Order Adjustment Settings")
            
            # Add explanation of priority rules
            with st.expander("📝 Priority Rules Explanation", expanded=False):
                st.markdown("""
                **Predicted_Order Computation Priority:**
                
                1. **Stock Subtraction**: `Predicted_Order = Predicted_Base_Quantity - Stock`
                2. **Box Adjustment**: If result is within ±tolerance of Box quantity, adjust to Box quantity
                3. **Scheme Adjustment**: If result is within ±tolerance of scheme total, apply scheme format
                
                **Examples:**
                - Base=14, Stock=4 → 14-4=10
                - If Box=10 and tolerance=2: 10 is exactly Box, so stays 10
                - If Scm=5+1 (total=6) and tolerance=2: |10-6|=4 > 2, so stays 10
                
                **Note**: Scheme results always sum to whole numbers (e.g., 4.5+0.5=5, not 4+0.5=4.5)
                """)
            
            # Create columns for compact layout
            col1, col2, col3, col4, col5 = st.columns([2, 1, 2, 1, 1])
            
            with col1:
                apply_box_adj = st.checkbox("Apply Box Quantity", value=True, key="apply_box_checkbox")
            
            with col2:
                box_tolerance_val = st.number_input("Box tolerance", min_value=0, max_value=10, value=2, step=1, key="box_tolerance_input")
            
            with col3:
                apply_scm_adj = st.checkbox("Apply Scm", value=True, key="apply_scm_checkbox")
            
            with col4:
                scm_tolerance_val = st.number_input("Scm tolerance", min_value=0, max_value=10, value=2, step=1, key="scm_tolerance_input")
            
            with col5:
                generate_button = st.button("Generate", type="primary", key="generate_button", help="Regenerate Predicted_Order with current settings")
            
            # Initialize session state for settings
            if 'current_settings' not in st.session_state:
                st.session_state.current_settings = {
                    'apply_box': True,
                    'box_tolerance': 2,
                    'apply_scm': True,
                    'scm_tolerance': 2
                }
            
            # Check if settings changed or Generate button clicked
            settings_changed = (
                st.session_state.current_settings['apply_box'] != apply_box_adj or
                st.session_state.current_settings['box_tolerance'] != box_tolerance_val or
                st.session_state.current_settings['apply_scm'] != apply_scm_adj or
                st.session_state.current_settings['scm_tolerance'] != scm_tolerance_val
            )
            
            # Regenerate if Generate button clicked or if this is initial load
            if generate_button or settings_changed or 'df_result_adjusted' not in st.session_state:
                # Update current settings
                st.session_state.current_settings = {
                    'apply_box': apply_box_adj,
                    'box_tolerance': box_tolerance_val,
                    'apply_scm': apply_scm_adj,
                    'scm_tolerance': scm_tolerance_val
                }
                
                # Reprocess with new settings
                with st.spinner("🔄 Regenerating predictions with new settings..."):
                    result_adjusted = process_uploaded_file(
                        uploaded_file, 
                        apply_box=apply_box_adj, 
                        box_tolerance=box_tolerance_val,
                        apply_scm=apply_scm_adj, 
                        scm_tolerance=scm_tolerance_val
                    )
                    if result_adjusted[0] is not None:
                        st.session_state.df_result_adjusted, st.session_state.df_processed_adjusted, st.session_state.styling_info_adjusted, st.session_state.expiry_styling_adjusted = result_adjusted
                        st.success("✅ Predictions regenerated successfully!")
                    else:
                        st.error("❌ Error regenerating predictions")
                        st.session_state.df_result_adjusted, st.session_state.df_processed_adjusted, st.session_state.styling_info_adjusted, st.session_state.expiry_styling_adjusted = None, None, None, None
            
            # Use adjusted results if available, otherwise fall back to original
            if 'df_result_adjusted' in st.session_state and st.session_state.df_result_adjusted is not None:
                df_result = st.session_state.df_result_adjusted
                df_processed = st.session_state.df_processed_adjusted
                styling_info = st.session_state.styling_info_adjusted
                expiry_styling = st.session_state.expiry_styling_adjusted
            
            st.markdown("---")  # Add separator line
            
            # Add new UI controls for filtering and editing
            st.markdown("### 🎛️ Display & Export Options")
            
            # Create columns for the filter controls
            filter_col1, filter_col2 = st.columns([1, 2])
            
            with filter_col1:
                ignore_no_order = st.checkbox(
                    "Ignore No Order", 
                    value=False, 
                    key="ignore_no_order_checkbox",
                    help="Hide rows where Predicted_Order = 'No Order' from display and export"
                )
            
            with filter_col2:
                # Get unique suppliers from the data
                supplier_col = None
                supplier_variations = ['Supplier', 'Supplier Name', 'supplier', 'supplier_name', 'SUPPLIER']
                
                for col_name in supplier_variations:
                    if col_name in df_result.columns:
                        supplier_col = col_name
                        break
                
                if supplier_col is None:
                    # Try case-insensitive search
                    for col in df_result.columns:
                        if 'supplier' in str(col).lower():
                            supplier_col = col
                            break
                
                if supplier_col and supplier_col in df_result.columns:
                    unique_suppliers = sorted(df_result[supplier_col].dropna().unique().tolist())
                    excluded_suppliers = st.multiselect(
                        "Exclude Suppliers",
                        options=unique_suppliers,
                        default=[],
                        key="excluded_suppliers_multiselect",
                        help="Select suppliers to exclude from display and export"
                    )
                else:
                    excluded_suppliers = []
                    st.info("No Supplier column found in data")
            
            # Determine the base data to filter (use edited data if available)
            base_data_for_filtering = df_result.copy()
            if 'edited_data' in st.session_state and st.session_state.edited_data is not None:
                # Use edited data for filtering, but ensure it has the same structure
                if (len(st.session_state.edited_data) == len(df_result) and 
                    list(st.session_state.edited_data.columns) == list(df_result.columns)):
                    base_data_for_filtering = st.session_state.edited_data.copy()
                    # Reset index to match original data for styling mapping
                    base_data_for_filtering.index = df_result.index
            
            # Apply filters to the data (now works with edited data)
            df_filtered = base_data_for_filtering.copy()
            
            # Filter out "No Order" rows if checkbox is selected
            if ignore_no_order:
                df_filtered = df_filtered[df_filtered['Predicted_Order'] != 'No Order']
            
            # Filter out excluded suppliers
            if supplier_col and excluded_suppliers:
                df_filtered = df_filtered[~df_filtered[supplier_col].isin(excluded_suppliers)]
            
            # Store original indices before resetting for styling mapping
            original_indices = df_filtered.index.tolist()
            
            # Reset index for filtered data to ensure proper display
            df_filtered = df_filtered.reset_index(drop=True)
            
            # Update styling info to match filtered data with new sequential indices
            filtered_styling_info = {}
            filtered_expiry_styling = {}
            
            if styling_info:
                # Map original indices to new sequential indices (0, 1, 2, ...)
                for new_idx, orig_idx in enumerate(original_indices):
                    if orig_idx in styling_info:
                        filtered_styling_info[new_idx] = styling_info[orig_idx]
            
            if expiry_styling:
                # Map original indices to new sequential indices (0, 1, 2, ...)
                for new_idx, orig_idx in enumerate(original_indices):
                    if orig_idx in expiry_styling:
                        filtered_expiry_styling[new_idx] = expiry_styling[orig_idx]
            
            st.markdown(f"**Showing {len(df_filtered)} of {len(base_data_for_filtering)} rows**")
            
            # Apply styling based on business rules
            def apply_business_rule_styling(df, styling_info, expiry_styling=None):
                styler = df.style
                
                # Apply business rule colors to Predicted_Order column
                if 'Predicted_Order' in df.columns:
                    for row_idx, style_info in styling_info.items():
                        if row_idx < len(df):
                            # Only apply background color if color is not None (skip "No Order" cases)
                            if style_info.get('color') is not None:
                                styler = styler.set_properties(
                                    subset=pd.IndexSlice[row_idx, 'Predicted_Order'],
                                    **{"background-color": style_info['color'], "color": "black"}
                                )
                
                # Apply expiry column highlighting
                if expiry_styling:
                    # Find expiry column
                    expiry_col = None
                    expiry_col_variations = ['Expiry', 'Expiry Date', 'Exp', 'Exp Date', 'expiry', 'expiry_date']
                    
                    for col_name in expiry_col_variations:
                        if col_name in df.columns:
                            expiry_col = col_name
                            break
                    
                    if expiry_col is None:
                        for col in df.columns:
                            if 'expiry' in str(col).lower() or 'exp' in str(col).lower():
                                expiry_col = col
                                break
                    
                    if expiry_col:
                        for row_idx, expiry_style in expiry_styling.items():
                            if row_idx < len(df):
                                styler = styler.set_properties(
                                    subset=pd.IndexSlice[row_idx, expiry_col],
                                    **{"background-color": expiry_style['color'], "color": "black"}
                                )
                
                # Highlight key columns with light background
                cols_lower_map = {c.lower(): c for c in df.columns}
                target_cols = ["name", "order", "predicted_order", "predicted_base_quantity"]
                highlight_cols = [cols_lower_map[c] for c in target_cols if c in cols_lower_map]
                
                if highlight_cols:
                    # Apply light background to key columns (but don't override business rule colors)
                    for col in highlight_cols:
                        if col != 'Predicted_Order':  # Don't override business rule styling
                            styler = styler.set_properties(
                                subset=[col],
                                **{"background-color": "#fce5cd", "color": "black"}
                            )
                
                return styler
            
            # Enhanced editable data grid with Excel preview features
            if 'Predicted_Order' in df_filtered.columns:
                st.markdown("### ✏️ Enhanced Editable Data Grid")
                st.info("💡 This grid shows exactly what your Excel export will look like - with colors, tooltips, and editable Predicted_Order values. Hover over cells for business rule explanations.")
                
                # The enhanced grid should display the filtered data directly
                # The filtering logic above already handles edited data properly
                initial_grid_data = df_filtered.copy()
                
                # Note: df_filtered already contains edited data if available,
                # and has been properly filtered by ignore_no_order and excluded_suppliers
                # No additional edit persistence logic needed here since df_filtered
                # is already built from base_data_for_filtering which includes edits
                
                # Create enhanced grid with color highlighting and tooltips
                try:
                    # For change tracking, we need to compare against the base data (before any user edits)
                    # If we have edited data in session state, use df_result as original
                    # If no edits yet, use the current data as both original and current
                    original_for_tracking = df_result if 'edited_data' in st.session_state and st.session_state.edited_data is not None else initial_grid_data
                    
                    # Create a unique key that changes when filters change to force grid refresh
                    grid_key = f"enhanced_editable_grid_{ignore_no_order}_{len(excluded_suppliers)}_{len(df_filtered)}"
                    
                    grid_response = create_enhanced_grid(
                        initial_grid_data,
                        styling_info=filtered_styling_info,
                        expiry_styling=filtered_expiry_styling,
                        original_data=original_for_tracking,  # Pass appropriate original data for change tracking
                        key=grid_key
                    )
                    
                    # Process grid changes and update session state
                    if grid_response and 'data' in grid_response:
                        # Process the changes from the enhanced grid
                        updated_df = process_grid_changes(
                            initial_grid_data,
                            grid_response,
                            product_key_columns=['Name', 'Supplier', 'Stock']
                        )
                        
                        # Apply the changes back to the full dataset (not just filtered data)
                        # This ensures we maintain the complete dataset with user edits
                        if 'edited_data' not in st.session_state or st.session_state.edited_data is None:
                            st.session_state.edited_data = df_result.copy()
                        
                        # Apply the changes from the grid to the full dataset using product key mapping
                        try:
                            # Create mapping of changes from the updated filtered data
                            changes_map = {}
                            for idx, row in updated_df.iterrows():
                                # Create product key
                                key_fields = []
                                if 'Name' in row:
                                    key_fields.append(str(row['Name']))
                                if 'Supplier' in row:
                                    key_fields.append(str(row['Supplier']))
                                if 'Stock' in row:
                                    key_fields.append(str(row['Stock']))
                                
                                if key_fields:
                                    product_key = '|'.join(key_fields)
                                    changes_map[product_key] = row['Predicted_Order']
                            
                            # Apply changes to the full dataset
                            for idx, row in st.session_state.edited_data.iterrows():
                                key_fields = []
                                if 'Name' in row:
                                    key_fields.append(str(row['Name']))
                                if 'Supplier' in row:
                                    key_fields.append(str(row['Supplier']))
                                if 'Stock' in row:
                                    key_fields.append(str(row['Stock']))
                                
                                if key_fields:
                                    product_key = '|'.join(key_fields)
                                    if product_key in changes_map:
                                        st.session_state.edited_data.iloc[idx, st.session_state.edited_data.columns.get_loc('Predicted_Order')] = changes_map[product_key]
                        
                        except Exception as e:
                            st.warning(f"Error applying grid changes: {str(e)}")
                            # Fallback: just store the updated filtered data
                            st.session_state.edited_data = updated_df
                        
                        # Show changes summary
                        show_grid_changes_summary(initial_grid_data, updated_df)
                        
                        # Also show overall changes from original data
                        if 'edited_data' in st.session_state and st.session_state.edited_data is not None:
                            # Count total changes across all data
                            total_changes = 0
                            try:
                                for idx, row in st.session_state.edited_data.iterrows():
                                    if idx < len(df_result):
                                        original_val = df_result.iloc[idx]['Predicted_Order']
                                        edited_val = row['Predicted_Order']
                                        if str(original_val) != str(edited_val):
                                            total_changes += 1
                                
                                if total_changes > 0:
                                    st.info(f"📊 Total changes made: {total_changes} rows edited across all data")
                            except Exception as e:
                                st.warning(f"Could not calculate total changes: {str(e)}")
                
                except Exception as e:
                    st.error(f"Error creating enhanced grid: {str(e)}")
                    st.warning("Falling back to basic editable grid...")
                    
                    # Fallback to basic st.data_editor
                    column_config = {
                        "Predicted_Order": st.column_config.TextColumn(
                            "Predicted_Order",
                            help="Edit predicted order values",
                            max_chars=20,
                        )
                    }
                    
                    edited_df = st.data_editor(
                        initial_grid_data,
                        column_config=column_config,
                        disabled=[col for col in initial_grid_data.columns if col != 'Predicted_Order'],
                        use_container_width=True,
                        key="fallback_editable_data_grid"
                    )
                    
                    st.session_state.edited_data = edited_df
            else:
                # Fallback if no Predicted_Order column - show enhanced grid in read-only mode
                try:
                    grid_response = create_enhanced_grid(
                        df_filtered,
                        styling_info=filtered_styling_info,
                        expiry_styling=filtered_expiry_styling,
                        key="readonly_enhanced_grid"
                    )
                except Exception as e:
                    st.warning(f"Could not create enhanced grid: {str(e)}")
                    st.dataframe(df_filtered, use_container_width=True)
                
            # Display tooltip information as expandable section
            if filtered_styling_info or filtered_expiry_styling:
                with st.expander("📝 View Highlighting Reasons", expanded=False):
                    st.markdown("**Rows with business rule highlighting:**")
                    tooltip_data = []
                    
                    # Combine Predicted_Order and Expiry tooltips
                    all_highlighted_rows = set(filtered_styling_info.keys())
                    if filtered_expiry_styling:
                        all_highlighted_rows.update(filtered_expiry_styling.keys())
                    
                    for row_idx in sorted(all_highlighted_rows):
                        if row_idx < len(df_filtered):
                            product_name = df_filtered.iloc[row_idx].get('Name', f'Row {row_idx + 1}')
                            predicted_order = df_filtered.iloc[row_idx].get('Predicted_Order', 'N/A')
                            
                            reasons = []
                            if row_idx in filtered_styling_info:
                                reasons.append(f"Predicted_Order: {filtered_styling_info[row_idx]['tooltip']}")
                            if filtered_expiry_styling and row_idx in filtered_expiry_styling:
                                reasons.append(f"Expiry: {filtered_expiry_styling[row_idx]['tooltip']}")
                            
                            tooltip_data.append({
                                'Product': str(product_name)[:50] + '...' if len(str(product_name)) > 50 else str(product_name),
                                'Predicted_Order': predicted_order,
                                'Reason': ' | '.join(reasons)
                            })
                    
                    if tooltip_data:
                        tooltip_df = pd.DataFrame(tooltip_data)
                        st.dataframe(tooltip_df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No business rules triggered for this dataset.")
                
                # Add legend for colors
                st.markdown("### 🎨 Color Legend")
                st.markdown("**Predicted_Order Column:**")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown('<div style="background-color: #ffcccc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🔴 Reddish: Days > 90 / Uneven Sales</div>', unsafe_allow_html=True)
                with col2:
                    st.markdown('<div style="background-color: #ccffcc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🟢 Greenish: Edited in the UI</div>', unsafe_allow_html=True)
                with col3:
                    st.markdown('<div style="background-color: #ffe6cc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🟠 Orangish: Box Adjustment ±2</div>', unsafe_allow_html=True)
                with col4:
                    st.markdown('<div style="background-color: #ffffcc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🟡 Yellowish: Low Customers ≤2</div>', unsafe_allow_html=True)
                
                st.markdown("**Expiry Column:**")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown('<div style="background-color: #ffcccc; padding: 5px; border-radius: 3px; text-align: center; color: black;">⚠️ URGENT: Expiring ≤ 1 month</div>', unsafe_allow_html=True)
                with col2:
                    st.markdown('<div style="background-color: #ffe6cc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🟠 Soon: Expiring ≤ 3 months</div>', unsafe_allow_html=True)
                with col3:
                    st.markdown('<div style="background-color: #fff2cc; padding: 5px; border-radius: 3px; text-align: center; color: black;">🟡 Moderate: Expiring ≤ 5 months</div>', unsafe_allow_html=True)

            # Download button
            st.markdown("### 💾 Download Results")
            
            # Show what will be exported (but don't generate Excel until download is clicked)
            # Determine which data to export - use edited data if available, otherwise filtered data
            export_df = df_filtered.copy()
            if 'edited_data' in st.session_state and st.session_state.edited_data is not None:
                # Apply the same filters to the edited data
                export_df = st.session_state.edited_data.copy()
                
                # Apply "Ignore No Order" filter to edited data
                if ignore_no_order:
                    export_df = export_df[export_df['Predicted_Order'] != 'No Order']
                
                # Apply supplier filter to edited data
                if supplier_col and excluded_suppliers:
                    export_df = export_df[~export_df[supplier_col].isin(excluded_suppliers)]
            
            # Show export summary
            export_summary_col1, export_summary_col2 = st.columns(2)
            with export_summary_col1:
                st.info(f"📊 **Export Preview:**\n- Total rows: {len(export_df)}\n- Filtered out: {len(base_data_for_filtering) - len(export_df)} rows")
            with export_summary_col2:
                if ignore_no_order:
                    st.info("🚫 'No Order' rows excluded")
                if excluded_suppliers:
                    st.info(f"🏢 {len(excluded_suppliers)} supplier(s) excluded")
                if 'edited_data' in st.session_state and st.session_state.edited_data is not None:
                    st.info("✏️ Manual edits included")
            
            # Generate Excel only when download button is clicked
            def generate_excel_with_styling():
                """Generate Excel file with styling only when needed."""
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='Predictions')
                    
                    # Apply styling to Excel file (only for filtered data)
                    if filtered_styling_info or filtered_expiry_styling:
                        from openpyxl.styles import PatternFill
                        from openpyxl.comments import Comment
                        
                        worksheet = writer.sheets['Predictions']
                        
                        # Find Predicted_Order column
                        predicted_order_col = None
                        for col_idx, col_name in enumerate(export_df.columns, 1):
                            if col_name == 'Predicted_Order':
                                predicted_order_col = col_idx
                                break
                        
                        # Find Expiry column
                        expiry_col = None
                        expiry_col_variations = ['Expiry', 'Expiry Date', 'Exp', 'Exp Date', 'expiry', 'expiry_date']
                        for col_idx, col_name in enumerate(export_df.columns, 1):
                            if col_name in expiry_col_variations or 'expiry' in str(col_name).lower():
                                expiry_col = col_idx
                                break
                        
                        # Apply Predicted_Order styling (adjusted for filtered data)
                        if filtered_styling_info and predicted_order_col:
                            for row_idx, style_info in filtered_styling_info.items():
                                excel_row = row_idx + 2  # +2 because Excel is 1-indexed and has header
                                if excel_row <= len(export_df) + 1:  # Ensure we don't exceed export data bounds
                                    cell = worksheet.cell(row=excel_row, column=predicted_order_col)
                                    
                                    # Only apply background color if color is not None (skip "No Order" cases)
                                    if style_info.get('color') is not None:
                                        color_hex = style_info['color'].replace('#', '')
                                        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                                        cell.fill = fill
                                    
                                    # Always add comment with tooltip (including "No Order" explanations)
                                    comment = Comment(style_info['tooltip'], 'Business Rules')
                                    cell.comment = comment
                        
                        # Apply Expiry column styling (adjusted for filtered data)
                        if filtered_expiry_styling and expiry_col:
                            for row_idx, expiry_style in filtered_expiry_styling.items():
                                excel_row = row_idx + 2  # +2 because Excel is 1-indexed and has header
                                if excel_row <= len(export_df) + 1:  # Ensure we don't exceed export data bounds
                                    cell = worksheet.cell(row=excel_row, column=expiry_col)
                                    
                                    # Apply background color
                                    color_hex = expiry_style['color'].replace('#', '')
                                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                                    cell.fill = fill
                                    
                                    # Add comment with expiry tooltip
                                    comment = Comment(expiry_style['tooltip'], 'Expiry Alert')
                                    cell.comment = comment
                
                excel_data = output.getvalue()
                
                # Generate filename with timestamp
                timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
                filename = f"pharmacy_predictions_{timestamp}.xlsx"
                
                # Save a copy to server for model retraining (only when download is clicked)
                try:
                    # Create reports directory if it doesn't exist
                    script_dir = Path(__file__).parent.absolute()
                    reports_dir = script_dir.parent / "data" / "reports"
                    reports_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Save server copy
                    server_file_path = reports_dir / filename
                    with open(server_file_path, 'wb') as f:
                        f.write(excel_data)
                    
                    st.success(f"✅ Server copy saved: {server_file_path.name}")
                    
                except Exception as e:
                    st.warning(f"⚠️ Could not save server copy: {str(e)}")
                
                return excel_data, filename
            
            # Single button that generates Excel and provides download automatically
            with st.spinner("Generating Excel file with styling..."):
                excel_data, filename = generate_excel_with_styling()
            
            # Provide download button (always available after generation)
            st.download_button(
                label="📅 Generate & Download Excel with Predictions",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                help="Click to download Excel file with all predictions, styling, and your edits. Server copy will be automatically saved."
            )
    
    # Footer
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align: center; padding: 2rem; color: rgba(255,255,255,0.7);">
        <p>🚀 Powered by Advanced Machine Learning | Built with Streamlit</p>
        <p>📊 Enhanced with XGBoost, Random Forest & Feature Engineering</p>
    </div>
    """, unsafe_allow_html=True)


def retraining_tab():
    """Model Retraining & Management functionality"""
    st.markdown("## 🔄 Model Retraining & Management")
    st.markdown("Upload modified Excel files to retrain the prediction model and manage model versions.")
    
    # Initialize model retrainer with correct path
    script_dir = Path(__file__).parent.absolute()
    models_dir = script_dir.parent / "models"
    retrainer = ModelRetrainer(str(models_dir))
    registry = ModelRegistry(str(models_dir))
    
    # Create columns for layout
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📎 Upload Training Data")
        
        # File uploader for multiple files
        uploaded_files = st.file_uploader(
            "Choose Excel files with modified Predicted_Order values",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Upload Excel files containing your corrected Predicted_Order values for retraining"
        )
        
        # Training options
        st.markdown("### ⚙️ Training Options")
        incremental_training = st.checkbox(
            "Incremental Training", 
            value=True, 
            help="Build upon the existing model (recommended for XGBoost)"
        )
        
        # Retrain button
        if st.button("🚀 Start Retraining", type="primary", disabled=not uploaded_files):
            if uploaded_files:
                with st.spinner("🔄 Retraining model... This may take a few minutes."):
                    result = retrainer.retrain_pipeline(uploaded_files, incremental=incremental_training)
                
                if result['success']:
                    st.session_state['retraining_result'] = result
                    st.success("✅ Model retraining completed successfully!")
                    st.rerun()
                else:
                    st.error(f"❌ Retraining failed: {result['error']}")
    
    with col2:
        st.markdown("### 📊 Current Model Status")
        
        # Display current active model info
        active_model = registry.get_active_model()
        if active_model:
            st.info(f"🏆 **Active Model:** {active_model['version']}")
            st.write(f"📅 **Trained:** {active_model['timestamp'][:10]}")
            
            if 'metrics' in active_model:
                metrics = active_model['metrics']
                st.write("📊 **Performance:**")
                col_a, col_b = st.columns(2)
                with col_a:
                    st.metric("RMSE", f"{metrics.get('rmse', 0):.3f}")
                    st.metric("R²", f"{metrics.get('r2', 0):.3f}")
                with col_b:
                    st.metric("Accuracy", f"{metrics.get('accuracy', 0):.3f}")
                    st.metric("F1 Score", f"{metrics.get('f1_score', 0):.3f}")
        else:
            st.warning("⚠️ No active model found")
    
    # Display retraining results if available
    if 'retraining_result' in st.session_state:
        result = st.session_state['retraining_result']
        
        st.markdown("---")
        st.markdown("### 🔍 Model Comparison")
        
        if result.get('comparison'):
            comparison = result['comparison']
            
            # Create comparison table
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("#### 📋 Current Model")
                current_metrics = comparison['current']
                for metric, value in current_metrics.items():
                    st.metric(metric.upper(), f"{value:.4f}")
            
            with col2:
                st.markdown("#### 🆕 New Model")
                new_metrics = comparison['new']
                for metric, value in new_metrics.items():
                    improvement = comparison['improvements'].get(metric, 0)
                    delta = f"{improvement:+.2f}%" if improvement != 0 else None
                    st.metric(metric.upper(), f"{value:.4f}", delta=delta)
            
            with col3:
                st.markdown("#### 🤖 Recommendation")
                recommendation = comparison['recommendation']
                if recommendation == 'use_new':
                    st.success("✅ **Use New Model**\n\nThe new model shows significant improvements.")
                else:
                    st.info("🔄 **Keep Current**\n\nCurrent model performs adequately.")
        
        # Model confirmation buttons
        st.markdown("### 📝 Finalize Model")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("✅ Accept New Model", type="primary"):
                with st.spinner("Finalizing new model..."):
                    final_version = retrainer.finalize_model(
                        result['temp_version'],
                        result['new_model_path'],
                        result['new_metrics']
                    )
                    st.success(f"✅ New model {final_version} is now active!")
                    del st.session_state['retraining_result']
                    st.rerun()
        
        with col2:
            if st.button("❌ Reject New Model"):
                # Clean up temporary model
                import os
                if os.path.exists(result['new_model_path']):
                    os.remove(result['new_model_path'])
                st.info("🗑️ New model rejected and cleaned up.")
                del st.session_state['retraining_result']
                st.rerun()
    
    # Model version management
    st.markdown("---")
    st.markdown("### 📚 Model Version Management")
    
    all_models = registry.get_all_models()
    if all_models:
        st.markdown("#### 📋 Available Model Versions")
        
        # Create a table of all models
        model_data = []
        for version, info in all_models.items():
            model_data.append({
                'Version': version,
                'Status': '🏆 Active' if info['is_active'] else '💾 Inactive',
                'Trained': info['timestamp'][:10],
                'RMSE': f"{info['metrics'].get('rmse', 0):.4f}" if 'metrics' in info else 'N/A',
                'R²': f"{info['metrics'].get('r2', 0):.4f}" if 'metrics' in info else 'N/A',
                'Accuracy': f"{info['metrics'].get('accuracy', 0):.4f}" if 'metrics' in info else 'N/A'
            })
        
        model_df = pd.DataFrame(model_data)
        st.dataframe(model_df, use_container_width=True, hide_index=True)
        
        # Rollback functionality
        st.markdown("#### ⏪ Rollback to Previous Version")
        
        inactive_models = {v: info for v, info in all_models.items() if not info['is_active']}
        if inactive_models:
            selected_version = st.selectbox(
                "Select version to rollback to:",
                options=list(inactive_models.keys()),
                format_func=lambda x: f"{x} (Trained: {inactive_models[x]['timestamp'][:10]})",
                help="Select a previous model version to make it active"
            )
            
            if selected_version:
                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("⏪ Rollback", type="secondary"):
                        with st.spinner(f"Rolling back to {selected_version}..."):
                            success = retrainer.rollback_to_version(selected_version)
                            if success:
                                st.success(f"✅ Successfully rolled back to {selected_version}!")
                                st.rerun()
                            else:
                                st.error(f"❌ Rollback to {selected_version} failed.")
                
                with col2:
                    if selected_version in inactive_models:
                        selected_info = inactive_models[selected_version]
                        if 'metrics' in selected_info:
                            st.write(f"📊 **{selected_version} Metrics:**")
                            metrics = selected_info['metrics']
                            metric_cols = st.columns(4)
                            with metric_cols[0]:
                                st.metric("RMSE", f"{metrics.get('rmse', 0):.4f}")
                            with metric_cols[1]:
                                st.metric("R²", f"{metrics.get('r2', 0):.4f}")
                            with metric_cols[2]:
                                st.metric("Accuracy", f"{metrics.get('accuracy', 0):.4f}")
                            with metric_cols[3]:
                                st.metric("F1", f"{metrics.get('f1_score', 0):.4f}")
        else:
            st.info("📚 Only one model version available. Train more models to enable rollback.")
    else:
        st.info("📚 No model versions found. Upload training data to create your first model.")
    
    # Help section
    with st.expander("📝 Help & Instructions", expanded=False):
        st.markdown("""
        **How to use Model Retraining:**
        
        1. **Upload Training Data**: Select Excel files containing corrected `Predicted_Order` values
        2. **Choose Training Type**: 
           - ✅ **Incremental**: Builds upon existing model (recommended)
           - ❌ **Full Retrain**: Trains completely new model
        3. **Review Comparison**: Compare new model metrics with current model
        4. **Accept or Reject**: Choose whether to use the new model
        5. **Rollback if Needed**: Switch back to previous versions anytime
        
        **Model Metrics Explained:**
        - **RMSE**: Root Mean Squared Error (lower is better)
        - **R²**: Coefficient of determination (higher is better, max 1.0)
        - **Accuracy**: Classification accuracy (higher is better)
        - **F1 Score**: Harmonic mean of precision and recall (higher is better)
        
        **Tips:**
        - Upload multiple files for better training data diversity
        - Incremental training preserves existing model knowledge
        - Keep track of model versions for easy rollback
        - Monitor metrics to ensure model improvements
        """)

if __name__ == "__main__":
    main()
